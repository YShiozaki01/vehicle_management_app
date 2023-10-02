import PySimpleGUI as sg
import sqlite3
import datetime
import pandas as pd
import openpyxl
from select_item import SelectItem, SelectVehicle
from transfer_abolition import TrfAbol
from application_form import PrtdataGen, PostingdataGen

sg.theme("SystemDefault")
DATABASE = "database/database.db"
SQL_DEPT = "SELECT code, name FROM M部署;"
SQL_SITUATION = "SELECT code, name FROM M状況;"
SQL_CARSIZE = "SELECT code, name FROM M車格;"
SQL_CARTYPE = "SELECT code, name FROM M種別;"
WB1 = "static/excel_template/運輸局申請様式.xlsm"
WS1_1 = "表紙"
WS2_1 = "別紙１"
WS3_1 = "別紙２"
WS4_1 = "別紙３"


# SQLのSELECT文で辞書型でレコードを取得するDB接続
def get_db_connection():
    connection = sqlite3.connect(DATABASE)
    connection.row_factory = sqlite3.Row
    return connection


# テーブルの最新ID番号を取得
def get_id_number(table_name):
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    sql1 = f"""
        SELECT * FROM {table_name};
        """
    cur.execute(sql1)
    if len(cur.fetchall()) == 0:
        new_id = 1
    else:
        sql = f"""
            SELECT max(id) FROM {table_name};
            """
        cur.execute(sql)
        result = cur.fetchone()
        new_id = result[0] + 1
    return new_id


# 車両3テーブルのデータを取得
def get_recd(
    table_name,
    cun="%",
):
    conn = get_db_connection()
    cur = conn.cursor()
    sql = f"""
        SELECT * FROM {table_name}
        WHERE company_use_number = '{cun}' and existence = 1;
        """
    cur.execute(sql)
    result = cur.fetchone()
    return result


# 入力フォームに再セットするデータを生成
def get_reset_data(company_use_number):
    conn = get_db_connection()
    cur = conn.cursor()
    sql = f"""
            SELECT
                a.company_use_number,
                a.classification,
                b.name as class_name,
                a.maker,
                a.model_number,
                a.body_number,
                a.model_year,
                a.car_size,
                c.name as size_name,
                a.load_capacity,
                a.specification,
                a.riding_capacity,
                a.fuel_type,
                d.department,
                d.implementation_date,
                d.circumstances,
                e.name as dept_name,
                f.basis_of_use,
                f.class_number,
                f.hiragana,
                f.designated_number
            FROM T車両台帳 as a
            LEFT JOIN M種別 as b on a.classification = b.code
            LEFT JOIN M車格 as c on a.car_size = c.code
            LEFT JOIN T車両履歴 as d
            on a.company_use_number = d.company_use_number
            LEFT JOIN M部署 as e on d.department = e.code
            LEFT JOIN T登録番号 as f
            on a.company_use_number = f.company_use_number
            WHERE a.company_use_number = '{company_use_number}'
            AND d.id = (SELECT max(id) FROM T車両履歴
            WHERE company_use_number = '{company_use_number}');
        """
    cur.execute(sql)
    result = cur.fetchone()
    return result


# 全入力フィールドをクリア
def clear_all():
    window["-correction-"].update(disabled=True)
    window["-transfer_abolition-"].update(disabled=True)
    window["-txt_abolition-"].update(visible=False)
    window["-in1-"].update("")
    window["-cd2-"].update("")
    window["-in3-"].update("")
    window["-in4-"].update("")
    window["-in5-"].update("")
    window["-in6-"].update("")
    window["-in7-"].update("")
    window["-cd3-"].update("")
    window["-in8-"].update("")
    window["-in9-"].update("")
    window["-in10-"].update("")
    window["-in11-"].update("")
    window["-in12-"].update("")
    window["-in1f-"].update("")
    window["-in2f-"].update("")
    window["-in3f-"].update("")
    window["-in4f-"].update("")
    window["-cd1-"].update("")
    window["-in2-"].update("")
    window["-in13-"].update("")
    window["-cd4-"].update("")
    # window["-chk_abolition-"].update(False)
    window["-correction-"].update(True)
    window["-in1-"].update(disabled=False)
    window["-in1-"].set_focus(True)


# 指定した社番で、実在の車両レコードが存在するか
def check_number(company_use_number):
    sql = f"""
        SELECT * FROM T車両台帳 WHERE company_use_number = '{company_use_number}';  
        """
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    cur.execute(sql)
    if len(cur.fetchall()) == 0:
        result = False
    else:
        result = True
    return result


# 3テーブルに入力データを挿入
def insert_data(signup):
    new_id1 = get_id_number("T車両台帳")
    new_id2 = get_id_number("T登録番号")
    new_id3 = get_id_number("T車両履歴")
    now_dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # T車両台帳テーブルに挿入するデータを生成
    car_info = []
    car_info.append(new_id1)  # T車両台帳<id>
    car_info.append(v["-in1-"])  # T車両台帳<company_use_number>
    car_info.append(v["-cd2-"])  # T車両台帳<classification>
    car_info.append(v["-in4-"])  # T車両台帳<maker>
    car_info.append(v["-in5-"])  # T車両台帳<model_number>
    car_info.append(v["-in6-"])  # T車両台帳<body_number>
    car_info.append(v["-in7-"])  # T車両台帳<model_year>
    car_info.append(v["-cd3-"])  # T車両台帳<car_size>
    car_info.append(v["-in9-"])  # T車両台帳<load_capacity>
    car_info.append(v["-in10-"])  # T車両台帳<specification>
    car_info.append(v["-in11-"])  # T車両台帳<riding_capacity>
    car_info.append(v["-in12-"])  # T車両台帳<fuel_type>
    car_info.append(1)  # T車両台帳<existence>
    car_info.append(now_dt)  # T車両台帳<update_date>
    # T登録番号テーブルに挿入するデータを生成
    reg_num = []
    reg_num.append(new_id2)  # T登録番号<id>
    reg_num.append(v["-in1-"])  # T登録番号<company_use_number>
    reg_num.append(v["-in1f-"])  # T登録番号<basis_of_use>
    reg_num.append(v["-in2f-"])  # T登録番号<class_number>
    reg_num.append(v["-in3f-"])  # T登録番号<hiragana>
    reg_num.append(v["-in4f-"])  # T登録番号<designated_number>
    reg_num.append(1)  # T登録番号<existence>
    reg_num.append(now_dt)  # T登録番号<update_date>
    # T車両履歴テーブルに挿入するデータを生成
    car_hist = []
    car_hist.append(new_id3)  # T車両履歴<id>
    car_hist.append(v["-in1-"])  # T車両履歴<company_use_number>
    car_hist.append(v["-cd1-"])  # T車両履歴<department>
    if signup:
        car_hist.append(v["-cd4-"])  # T車両履歴<circumstances>（修正）
    else:
        car_hist.append("A")  # T車両履歴<circumstances>（新規）
    car_hist.append(v["-in13-"])  # T車両履歴<implementation_date>
    car_hist.append(1)  # T車両履歴<existence>
    car_hist.append(now_dt)  # T車両履歴<update_date>
    # テーブルにデータを挿入
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    sql1 = """INSERT INTO T車両台帳
            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);"""
    sql2 = "INSERT INTO T登録番号 VALUES(?, ?, ?, ?, ?, ?, ?, ?);"
    sql3 = "INSERT INTO T車両履歴 VALUES(?, ?, ?, ?, ?, ?, ?);"
    cur.execute(sql1, car_info)
    cur.execute(sql2, reg_num)
    cur.execute(sql3, car_hist)
    conn.commit()


# T車両履歴テーブルにcircumstancesがA以外のレコードがあるか
def check_circumstances(company_use_number):
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    sql = f"""
        SELECT * FROM T車両履歴 WHERE company_use_number = '{company_use_number}';
        """
    cur.execute(sql)
    number = len(cur.fetchall())
    return number


# テーブルからレコードを削除
def delete_data(company_use_number):
    sql = f"""
        DELETE FROM T車両台帳 WHERE company_use_number = '{company_use_number}'
        and existence = 1;
        DELETE FROM T車両履歴 WHERE company_use_number = '{company_use_number}'
        and existence = 1;
        DELETE FROM T登録番号 WHERE company_use_number = '{company_use_number}'
        and existence = 1;
        """
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    cur.executescript(sql)
    conn.commit()

# 登録番号以外の入力フィールドを使用不可（1）
def set_disabled_1():
    window["-btn1-"].update(disabled=True)
    window["-btn2-"].update(disabled=True)
    window["-in4-"].update(disabled=True)
    window["-in5-"].update(disabled=True)
    window["-in6-"].update(disabled=True)
    window["-in7-"].update(disabled=True)
    window["-btn3-"].update(disabled=True)
    window["-in9-"].update(disabled=True)
    window["-in10-"].update(disabled=True)
    window["-in11-"].update(disabled=True)
    window["-in12-"].update(disabled=True)
    window["-in13-"].update(disabled=True)


# （1）を解除
def reset_disabled_1():
    window["-btn1-"].update(disabled=False)
    window["-btn2-"].update(disabled=False)
    window["-in4-"].update(disabled=False)
    window["-in5-"].update(disabled=False)
    window["-in6-"].update(disabled=False)
    window["-in7-"].update(disabled=False)
    window["-btn3-"].update(disabled=False)
    window["-in9-"].update(disabled=False)
    window["-in10-"].update(disabled=False)
    window["-in11-"].update(disabled=False)
    window["-in12-"].update(disabled=False)
    window["-in13-"].update(disabled=False)


# T登録番号を更新
def update_regnum(company_use_number, num1, num2, num3, num4):
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    sql = f"""
        UPDATE T登録番号 SET
            basis_of_use = '{num1}',
            class_number = '{num2}',
            hiragana = '{num3}',
            designated_number = '{num4}'
        WHERE company_use_number = '{company_use_number}';
        """
    cur.execute(sql)
    conn.commit()


# 8桁数字を日付に変換
def change_date(value):
    # 入力された値が8桁数字かチェック（数値ではなく整数文字か判定）
    target = value
    if target.isdigit() and len(target) == 8:
        date = pd.to_datetime(target)
        format_date = format(date, "%Y/%m/%d")
    else:
        format_date = ""
    return format_date


# フィールドの使用不可を解除
def release_disabled(pattern):
    window["-correction-"].update(disabled=False)
    window["-transfer_abolition-"].update(disabled=False)
    window["-btn_prtadd-"].update(disabled=False)


# フィールドの使用不可を再設定
def reset_disabled():
    window["-correction-"].update(disabled=True)
    window["-transfer_abolition-"].update(disabled=True)
    window["-btn_prtadd-"].update(disabled=True)


# TW系テーブルをクリア
def clear_tw():
    conn = sqlite3.connect("database/database.db")
    cur = conn.cursor()
    sql = """
        DELETE FROM TPrintWork;
        DELETE FROM TW内訳別申請台数;
        DELETE FROM TW内訳別保有台数;
        DELETE FROM TW車格別保有台数;
        DELETE FROM TW申請車両;
        """
    cur.executescript(sql)
    conn.commit()

# ウインドウレイアウト
prt_frame_layout = [
    [
        sg.B(
            "申請車両登録",
            k="-btn_prtadd-",
            size=(15, 0),
            font=("Yu Gothic UI", 8),
            disabled=True,
        ),
        sg.Push(),
        sg.B(
            "作成", k="-btn_print-", size=(15, 0), font=("Yu Gothic UI", 8),
            disabled=True
        ),
    ]
]
frame_layout = [
    [
        sg.T("使用の本拠", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in1f-", size=(10, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("分類番号", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in2f-", size=(10, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("ひらがな", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in3f-", size=(5, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("指定番号", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in4f-", size=(10, 0), font=("Yu Gothic UI", 8)),
    ],
]
layout = [
    [sg.T("車両入力", font=("Yu Gothic UI", 11))],
    [
        sg.T("社番", size=(10, 0), font=("Yu Gothic UI", 8),
             text_color="#FF0000"),
        sg.I(k="-in1-", size=(15, 0), font=("Yu Gothic UI", 8)),
        sg.T("廃止登録済み", k="-txt_abolition-", background_color="#FF0000",
             text_color="#ffffff", visible=False),
    ],
    [
        sg.T("部署", size=(10, 0), font=("Yu Gothic UI", 8),
             text_color="#FF0000"),
        sg.B("検索", k="-btn1-", font=("Yu Gothic UI", 8)),
        sg.I(k="-in2-", font=("Yu Gothic UI", 8), size=(20, 0), readonly=True),
        sg.I(k="-cd1-", font=("Yu Gothic UI", 8), size=(4, 0), visible=False),
    ],
    [
        sg.T("内訳", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.B("検索", k="-btn2-", font=("Yu Gothic UI", 8)),
        sg.I(k="-in3-", font=("Yu Gothic UI", 8), size=(20, 0), readonly=True),
        sg.I(k="-cd2-", font=("Yu Gothic UI", 8), size=(4, 0), visible=False),
    ],
    [
        sg.T("車名", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in4-", size=(20, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("型式", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in5-", size=(30, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("車体番号", size=(10, 0), font=("Yu Gothic UI", 8),
             text_color="#FF0000"),
        sg.I(k="-in6-", size=(30, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("年式", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in7-", size=(20, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("車格", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.B("検索", k="-btn3-", font=("Yu Gothic UI", 8)),
        sg.I(k="-in8-", font=("Yu Gothic UI", 8), size=(20, 0), readonly=True),
        sg.I(k="-cd3-", font=("Yu Gothic UI", 8), size=(4, 0), visible=False),
    ],
    [
        sg.T("最大積載量", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in9-", size=(10, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("形状", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in10-", size=(30, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("乗車定員", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in11-", size=(10, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("燃料種類", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in12-", size=(20, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.T("実施日", size=(10, 0), font=("Yu Gothic UI", 8)),
        sg.I(k="-in13-", size=(20, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.Frame(title="登録番号", font=("Yu Gothic UI", 8), layout=frame_layout),
        sg.I(k="-cd4-", font=("Yu Gothic UI", 8), size=(4, 0), visible=False),
        sg.Push(),
        sg.B("検索", k="-btn_search-", size=(6, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.B("取消", k="-btn_cancel-", size=(6, 0), font=("Yu Gothic UI", 8)),
        sg.Push(),
        sg.Radio(
            "修正",
            group_id="process1",
            font=("Yu Gothic UI", 8),
            key="-correction-",
            disabled=True,
            default=True,
        ),
        sg.Radio(
            "移動・廃止",
            group_id="process1",
            font=("Yu Gothic UI", 8),
            key="-transfer_abolition-",
            disabled=True,
        ),
        sg.B("登録", k="-btn_register-", size=(6, 0), font=("Yu Gothic UI", 8)),
    ],
    [
        sg.Frame(
            title="申請書作成",
            font=("Yu Gothic UI", 8),
            size=(290, 50),
            layout=prt_frame_layout,
        )
    ],
]
window = sg.Window(
    "車両入力", layout, font=("Yu Gothic UI", 8), size=(310, 560),
    disable_close=False
)
window.finalize()

# エンターキー押下
window["-in1-"].bind("<Return>", "_r")
window["-btn1-"].bind("<Return>", "_r")
window["-btn2-"].bind("<Return>", "_r")
window["-in4-"].bind("<Return>", "_r")
window["-in5-"].bind("<Return>", "_r")
window["-in6-"].bind("<Return>", "_r")
window["-in7-"].bind("<Return>", "_r")
window["-btn3-"].bind("<Return>", "_r")
window["-in9-"].bind("<Return>", "_r")
window["-in10-"].bind("<Return>", "_r")
window["-in11-"].bind("<Return>", "_r")
window["-in12-"].bind("<Return>", "_r")
window["-in13-"].bind("<Return>", "_r")
window["-in1f-"].bind("<Return>", "_r")
window["-in2f-"].bind("<Return>", "_r")
window["-in3f-"].bind("<Return>", "_r")
window["-in4f-"].bind("<Return>", "_r")
window["-in1-"].SetFocus(True)

selection_mode = False
while True:
    e, v = window.read()
    if e == "-in1-_r":
        window["-btn1-"].set_focus(True)
    if e == "-btn1-_r":
        window["-btn2-"].set_focus(True)
    if e == "-btn2-_r":
        window["-in4-"].set_focus(True)
    if e == "-in4-_r":
        window["-in5-"].set_focus(True)
    if e == "-in5-_r":
        window["-in6-"].set_focus(True)
    if e == "-in6-_r":
        window["-in7-"].set_focus(True)
    if e == "-in7-_r":
        window["-btn3-"].set_focus(True)
    if e == "-btn3-_r":
        window["-in9-"].set_focus(True)
    if e == "-in9-_r":
        window["-in10-"].set_focus(True)
    if e == "-in10-_r":
        window["-in11-"].set_focus(True)
    if e == "-in11-_r":
        window["-in12-"].set_focus(True)
    if e == "-in12-_r":
        window["-in13-"].set_focus(True)
    if e == "-in13-_r":
        strdate = change_date(v["-in13-"])
        if strdate == "":
            sg.popup("8桁数字で入力してください", font=("Yu Gothic UI", 8))
        else:
            window["-in13-"].update(strdate)
            window["-in1f-"].set_focus(True)
    if e == "-in1f-_r":
        window["-in2f-"].set_focus(True)
    if e == "-in2f-_r":
        window["-in3f-"].set_focus(True)
    if e == "-in3f-_r":
        window["-in4f-"].set_focus(True)
    if e == "-in4f-_r":
        window["-btn_register-"].set_focus(True)
    if e == "-btn1-":
        si1 = SelectItem(SQL_DEPT)
        items = si1.open_sub_window()
        if items:
            window["-in2-"].update(items[1])
            window["-cd1-"].update(items[0])
    if e == "-btn2-":
        si2 = SelectItem(SQL_CARTYPE)
        items = si2.open_sub_window()
        if items:
            window["-in3-"].update(items[1])
            window["-cd2-"].update(items[0])
    if e == "-btn3-":
        si3 = SelectItem(SQL_CARSIZE)
        items = si3.open_sub_window()
        if items:
            window["-in8-"].update(items[1])
            window["-cd3-"].update(items[0])
    if e == "-btn_register-":
        if not selection_mode:
            reality = check_number(v["-in1-"])
            if reality:
                sg.popup("登録済みの車番です。", title="確認", font=("Yu Gothic UI", 8))
            else:
                insert_data(selection_mode)
        if selection_mode:
            if v["-correction-"]:
                if check_circumstances(v["-in1-"]) == 1:
                    delete_data(v["-in1-"])
                    insert_data(selection_mode)
                else:
                    update_regnum(v["-in1-"], v["-in1f-"], v["-in2f-"], v["-in3f-"], v["-in4f-"])
                    reset_disabled_1()
            elif v["-transfer_abolition-"]:
                ta = TrfAbol(v["-in1-"], v["-in2-"], v["-cd1-"])
                ta.open_sub_window2()
            selection_mode = False
            reset_disabled()
        clear_all()
    if e == "-btn_search-":
        selection_mode = True
        company_use_number = v["-in1-"] if v["-in1-"] else "%"
        body_number = v["-in6-"] if v["-in6-"] else "%"
        department = v["-cd1-"] if v["-cd1-"] else "%"
        sv = SelectVehicle(
            company_use_number, body_number, department
        )
        vinfo = sv.open_sub_window()
        if vinfo:
            window["-in1-"].update(disabled=True)
            release_disabled(check_circumstances(vinfo[0]))
            record = get_reset_data(vinfo[0])
            if record["circumstances"] == "D":
                window["-txt_abolition-"].update(visible=True)
            window["-in1-"].update(record["company_use_number"])
            window["-cd2-"].update(record["classification"])
            window["-in3-"].update(record["class_name"])
            window["-in4-"].update(record["maker"])
            window["-in5-"].update(record["model_number"])
            window["-in6-"].update(record["body_number"])
            window["-in7-"].update(record["model_year"])
            window["-cd3-"].update(record["car_size"])
            window["-in8-"].update(record["size_name"])
            window["-in9-"].update(record["load_capacity"])
            window["-in10-"].update(record["specification"])
            window["-in11-"].update(record["riding_capacity"])
            window["-in12-"].update(record["fuel_type"])
            window["-in1f-"].update(record["basis_of_use"])
            window["-in2f-"].update(record["class_number"])
            window["-in3f-"].update(record["hiragana"])
            window["-in4f-"].update(record["designated_number"])
            window["-cd1-"].update(record["department"])
            window["-in2-"].update(record["dept_name"])
            window["-in13-"].update(record["implementation_date"])
            window["-cd4-"].update(record["circumstances"])
            if check_circumstances(vinfo[0]) != 1:
                set_disabled_1()
    if e == "-btn_cancel-":
        clear_all()
        reset_disabled()
    if e == "-btn_prtadd-":
        window["-btn_print-"].update(disabled=False)
        prt_conditions = [v["-in1-"], v["-in13-"], v["-cd1-"], v["-cd4-"]]
        prtg = PrtdataGen(prt_conditions)
        # 申請車両抽出条件リストを作成
        prtg.insert_pw()
        # 申請車両情報をTW申請車両に書き出し
        prtg.application_vehicle()
        reset_disabled_1()
        reset_disabled()
        clear_all()
    if e == "-btn_print-":
        pstg = PostingdataGen()
        list = pstg.get_appl_info()
        answ = sg.popup_ok_cancel(f"""以下の申請書を作成します。よろしいですか？\n\n{list}""",
                                  title="確認", font=("Yu Gothic UI", 8))
        if answ == "OK":
            # 申請前後の台数差分を取得してテーブルに書き出し
            pstg.get_difference()
            print_list = pstg.get_print_list()
            for dept_code, impl_date, dept_name in print_list:
                # 部署別内訳別保有台数を取得してテーブルに書き出し
                pstg.number_of_class(impl_date)
                # 部署別車格別保有台数を取得してテーブルに書き出し
                pstg.number_of_size(impl_date)
                # 申請書の別紙1「種別（普通車）」表に転記用のデータを作成
                wb = openpyxl.load_workbook(WB1)
                ws1 = wb[WS1_1]
                ws2 = wb[WS2_1]
                ws3 = wb[WS3_1]
                ws4 = wb[WS4_1]
                sql = f"""
                    SELECT breakdown_dept, implementation_date FROM TW内訳別申請台数
                    WHERE application_dept = '{dept_code}'
                    AND implementation_date = '{impl_date}'
                    GROUP by breakdown_dept, implementation_date;
                    """
                conn = sqlite3.connect(DATABASE)
                cur = conn.cursor()
                cur.execute(sql)
                dept_b_list = cur.fetchall()
                r = 10
                for dept_b in dept_b_list:
                    sql = f"""
                        SELECT classification, number_of_units FROM TW内訳別保有台数
                        WHERE department = '{dept_b[0]}';
                        """
                    conn = sqlite3.connect(DATABASE)
                    cur = conn.cursor()
                    cur.execute(sql)
                    units = dict(cur.fetchall())
                    sql = f"""
                        SELECT classification, adjustment FROM TW内訳別申請台数
                        WHERE application_dept = '{dept_code}'
                        AND breakdown_dept = '{dept_b[0]}'
                        AND implementation_date = '{dept_b[1]}';
                        """
                    conn = sqlite3.connect(DATABASE)
                    cur = conn.cursor()
                    cur.execute(sql)
                    diff = dict(cur.fetchall())
                    sql = f"""
                        SELECT official_use_name FROM M部署
                        WHERE code = '{dept_b[0]}';
                        """
                    conn = sqlite3.connect(DATABASE)
                    cur = conn.cursor()
                    cur.execute(sql)
                    name = cur.fetchone()[0]
                    ws2.cell(r, 1).value = name
                    c = 6
                    for class_code in ["C1", "C2", "C3", "C4"]:
                        val_a = units[class_code] if class_code in units else 0
                        val_b = (units[class_code] if class_code in units else 0) + (diff[class_code] if class_code in diff else 0)
                        ws2.cell(r, c).value = val_a
                        ws2.cell(r, c + 10).value = val_b
                        c += 2
                    r += 2
                # 申請書の別紙1「増減車両の明細」表に転記するデータを作成して転記
                posting_data = pstg.gen_posting_data2(ws2, dept_code, impl_date)
                # 申請書の別紙2「自動車倉庫の位置及び収容能力並びに必要面積」に転記するデータを作成
                # 作成した申請書Excelファイルを保存
                strdate = dept_b[1].replace("/", "-")
                wb.save(f"{dept_name}_{strdate}.xlsx")
            pstg.clear_tw()
            window["-btn_print-"].update(disabled=True)
            sg.popup("Excel「申請書」を作成しました。", title="作成完了", font=("Yu Gothic UI", 8))
        else:
            sg.popup("申請書作成を中止しました。", title="作成中止", font=("Yu Gothic UI", 8))
        clear_tw()
    if e == None:
        break
window.close()
