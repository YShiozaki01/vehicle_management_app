import PySimpleGUI as sg
import sqlite3
import openpyxl 

sg.theme("SystemDefault")
DATABASE = "database/database.db"


# データベースに接続
def get_db_connection():
    connection = sqlite3.connect(DATABASE)
    connection.row_factory = sqlite3.Row
    return connection


class PrtdataGen:
    def __init__(self, prt_conditions):
        self.prt_conditions = prt_conditions
        self.company_use_number = prt_conditions[0]
        self.implementation_date = prt_conditions[1]
        self.department = prt_conditions[2]
        self.circumstances = prt_conditions[3]

    # 部署別保有台数
    def number_of_dept(self):
        sql = f"""
            SELECT sum(existence) as number, department FROM T車両履歴
            GROUP by department
            WHERE department = '{self.department}
            AND implementation_date <= '{self.implementation_date}';'
            """
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(sql)
        result = cur.fetchall()
        return result

    # 申請車両情報
    def application_vehicle(self):
        if self.circumstances == "T":
            dept_org = self.get_dept_org()
            dept_org_code = dept_org[2]
            dept_org_name = dept_org[3]
        else:
            dept_org_code = ""
            dept_org_name = ""
        incr_decr = self.get_incr_decr()
        sql = f"""
            INSERT INTO TW申請車両 
            SELECT
            a.company_use_number,
            b.circumstances,
            b.department,
            c.official_use_name,
            a.classification,
            d.name as class_name,
            a.maker,
            a.model_year,
            a.load_capacity,
            a.body_number,
            e.basis_of_use || class_number || hiragana || designated_number
            as reg_no,
            a.specification,
            '{dept_org_code}',
            '{dept_org_name}',
            a.riding_capacity,
            a.existence,
            '{incr_decr}',
            b.implementation_date
            FROM T車両台帳 as a
            LEFT JOIN T車両履歴 as b
            on a.company_use_number = b.company_use_number
            LEFT JOIN M部署 as c on b.department = c.code
            LEFT JOIN M種別 as d on a.classification = d.code
            LEFT JOIN T登録番号 as e
            on a.company_use_number = e.company_use_number
            WHERE a.company_use_number = '{self.company_use_number}'
            AND b.implementation_date = '{self.implementation_date}'
            AND b.id = (SELECT max(id) FROM T車両履歴
            WHERE company_use_number = '{self.company_use_number}'
            AND implementation_date = '{self.implementation_date}')
            AND e.id = (SELECT max(id) FROM T登録番号
            WHERE company_use_number = '{self.company_use_number}');"""
        conn = sqlite3.connect(DATABASE)
        cur = conn.cursor()
        cur.execute(sql)
        conn.commit()

    # 増減区分を取得
    def get_incr_decr(self):
        if self.circumstances in ("A", "T"):
            incr_decr = "I"
        elif self.circumstances in ("D", "F"):
            incr_decr = "D"
        return incr_decr

    # 直近の移動元履歴を取得
    def get_dept_org(self):
        sql = f"""
            SELECT a.id, a.company_use_number, a.department,
            b.official_use_name as dept_name, a.circumstances
            FROM T車両履歴 as a
            LEFT JOIN M部署 as b on a.department = b.code
            WHERE a.id =
            (SELECT max(id) FROM T車両履歴
            WHERE implementation_date <= '{self.implementation_date}'
            AND company_use_number = '{self.company_use_number}'
            AND circumstances = 'F');
            """
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(sql)
        result = cur.fetchone()
        return result

    # TPrintWorkへ作成条件を挿入
    def insert_pw(self):
        sql = """
            INSERT INTO TPrintWork VALUES(?, ?, ?, ?);
            """
        conn = sqlite3.connect(DATABASE)
        cur = conn.cursor()
        cur.execute(sql, self.prt_conditions)
        conn.commit()


class PostingdataGen:
    def __init__(self):
        pass

    # 申請車両の社番と申請区分のリスト
    def get_appl_info(self):
        sql = """
            SELECT company_use_number,
            CASE
            WHEN prt_number = 'A' THEN '増車'
            WHEN prt_number = 'D' THEN '減車'
            WHEN prt_number = 'T' THEN '営配'
            end
            as division
            FROM TPrintWork;
            """
        conn = sqlite3.connect(DATABASE)
        cur = conn.cursor()
        cur.execute(sql)
        result = cur.fetchall()
        list = ""
        for num, div in result:
            row = f"{num}　{div}" + "\n"
            list += row
        return list

    # 申請営業所のリストを取得
    def get_print_list(self):
        sql = """SELECT department, implementation_date
            , min(official_use_name) as oun
            FROM TW申請車両
            GROUP by department, implementation_date;
            """
        conn = sqlite3.connect(DATABASE)
        cur = conn.cursor()
        cur.execute(sql)
        result = cur.fetchall()
        dept_list = []
        for dept in result:
            dept_list.append(dept)
        return dept_list

    # 部署別内訳別保有台数
    def number_of_class(self, implementation_date):
        sql = f"""
            INSERT INTO TW内訳別保有台数
            SELECT department, classification, sum(existence) as number FROM
            (SELECT a.company_use_number, b.classification, a.department,
            a.existence
            FROM (SELECT * FROM T車両履歴
            WHERE implementation_date <= '{implementation_date}') as a
            LEFT JOIN T車両台帳 as b
            on a.company_use_number = b.company_use_number)
            GROUP by department, classification;
            """
        conn = sqlite3.connect(DATABASE)
        cur = conn.cursor()
        cur.execute(sql)
        conn.commit()

    # 部署別車格別保有台数
    def number_of_size(self, implementation_date):
        sql = f"""
            INSERT INTO TW車格別保有台数
            SELECT department, car_size, sum(existence) as number FROM
            (SELECT a.company_use_number, b.car_size, a.department, a.existence
            FROM (SELECT * FROM T車両履歴 WHERE existence = 1
            AND implementation_date <= '{implementation_date}') as a
            LEFT JOIN T車両台帳 as b
            on a.company_use_number = b.company_use_number)
            GROUP by department, car_size;
            """
        conn = sqlite3.connect(DATABASE)
        cur = conn.cursor()
        cur.execute(sql)
        conn.commit()

    # 申請前後の差分
    def get_difference(self):
        sql = """
            INSERT INTO TW内訳別申請台数
            SELECT dpt_a, dpt_b, classification, sum(adjustment),
            implementation_date as adj FROM ( 
            SELECT department as dpt_a, department as dpt_b, classification,
            incr_decr,
            CASE
            WHEN incr_decr = 'I' THEN -1
            WHEN incr_decr = 'D' THEN 1
            end
            as adjustment,
            implementation_date
            FROM TW申請車両)
            GROUP by dpt_a, dpt_b, classification, implementation_date;
            INSERT INTO TW内訳別申請台数
            SELECT dpt_a, dpt_b, classification, sum(adjustment),
            implementation_date as adj FROM (
            SELECT department as dpt_a, dept_org as dpt_b, classification,
            incr_decr,
            1 as adjustment,
            implementation_date
            FROM TW申請車両
            WHERE not dept_org = "")
            GROUP by dpt_a, dpt_b, classification, implementation_date;
            """
        conn = sqlite3.connect(DATABASE)
        cur = conn.cursor()
        cur.executescript(sql)
        conn.commit()

    # TW内訳別保有台数とTW車格別保有台数をクリア
    def clear_tw(self):
        sql = """
            DELETE FROM TW内訳別保有台数;
            DELETE FROM TW車格別保有台数;
            """
        conn = sqlite3.connect(DATABASE)
        cur = conn.cursor()
        cur.executescript(sql)
        conn.commit()

    # 別紙1「増減車両の明細」に転記するデータを生成
    def gen_posting_data2(self, ws, department, implementation_date):
        sql = f"""
            SELECT
                CASE
                    WHEN incr_decr = "I" THEN "増車"
                    WHEN incr_decr = "D" THEN "減車"
                END as division,
                official_use_name,
                class_name,
                maker,
                model_year,
                load_capacity,
                body_number,
                reg_no,
                specification,
                dept_org_name
            FROM TW申請車両
            WHERE department = '{department}'
            AND implementation_date = '{implementation_date}';
            """
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(sql)
        result = list(cur.fetchall())
        r = 25
        for row in result:
            ws.cell(r, 1).value = row["division"]
            ws.cell(r, 3).value = row["official_use_name"]
            ws.cell(r, 6).value = row["class_name"]
            ws.cell(r, 8).value = row["maker"]
            i = row["model_year"].find("年")
            ws.cell(r, 11).value = row["model_year"][:i + 1]
            ws.cell(r, 13).value = row["load_capacity"]
            ws.cell(r, 16).value = row["body_number"]
            ws.cell(r + 1, 16).value = row["reg_no"]
            ws.cell(r, 20).value = row["specification"]
            ws.cell(r, 23).value = row["dept_org_name"]
            r += 2

    # 別紙2「自動車倉庫の位置及び収容能力並びに必要面積」に転記するデータを生成
    def gen_posting_data3(self, ws, department, implementation_date):
        sql = f"""
            SELECT b.car_size, sum(a.existence) as number
            FROM T車両履歴 as a
            LEFT JOIN T車両台帳 as b
            on a.company_use_number = b.company_use_number
            WHERE a.implementation_date <= '{implementation_date}'
            AND a.department = '{department}'
            GROUP by b.car_size;
            """
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(sql)
        result = cur.fetchall()
        for row in result:
            if row["car_size"] == "1":
                ws["P5"] = row["number"]
            elif row["car_size"] == "2":
                ws["P7"] = row["number"]
            elif row["car_size"] == "3":
                ws["P9"] = row["number"]
            elif row["car_size"] == "4":
                ws["P11"] = row["number"]

    # 別紙2「自動車倉庫の位置及び収容能力並びに必要面積」に転記する駐車場情報を抽出
    def get_garage(self, ws, department):
        sql = f"""
            SELECT b.official_use_name, a.address, a.capacity FROM M駐車場 as a
            LEFT JOIN M部署 as b on a.code = b.code
            WHERE a.code = '{department}' ORDER by a.row_number;
            """
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(sql)
        result = cur.fetchall()
        r = 5
        for row in result:
            ws[f"A{r}"] = row["official_use_name"]
            ws[f"F{r}"] = row["address"]
            ws[f"M{r}"] = row["capacity"]
            r += 2

    # 別紙1「種別（普通車）」に転記する運行車データを生成
    def get_service_vehicle(self, ws, department, r):
        conn = get_db_connection()
        cur = conn.cursor()
        sql = f"SELECT * FROM M運行車 WHERE code = '{department}';"
        cur.execute(sql)
        result = cur.fetchall()
        ws[f"F{r + 1}"] = 0
        ws[f"H{r + 1}"] = 0
        ws[f"J{r + 1}"] = 0
        ws[f"L{r + 1}"] = 0
        ws[f"P{r + 1}"] = 0
        ws[f"R{r + 1}"] = 0
        ws[f"T{r + 1}"] = 0
        ws[f"V{r + 1}"] = 0
        position_new = {"C1": "F", "C2": "H", "C3": "J", "C4": "L"}
        position_old = {"C1": "P", "C2": "R", "C3": "T", "C4": "V"}
        for row in result:
            ws[f"{position_new[row['classification']]}{r + 1}"] = row["nov"]
            ws[f"{position_old[row['classification']]}{r + 1}"] = row["nov"]

    # 別紙3「運行管理者、整備管理者の選任状況」に転記するデータを生成
    def get_practitioners(self, ws, department):
        conn = get_db_connection()
        cur = conn.cursor()
        sql1 = f"""
            SELECT b.official_use_name, a.name, a.date_of_acquisition
            FROM M実務者 as a
            LEFT JOIN M部署 as b on a.code = b.code
            WHERE a.code = '{department}'
            AND division = 'F' ORDER by row_number;
            """
        sql2 = f"""
            SELECT b.official_use_name, a.name, a.date_of_acquisition
            FROM M実務者 as a
            LEFT JOIN M部署 as b on a.code = b.code
            WHERE a.code = '{department}'
            AND division = 'M' ORDER by row_number;
            """
        cur.execute(sql1)
        result = cur.fetchall()
        r = 5
        for row in result:
            ws[f"A{r}"] = row["official_use_name"]
            ws[f"F{r}"] = row["name"]
            ws[f"K{r}"] = row["date_of_acquisition"]
            r += 2
        cur.execute(sql2)
        result = cur.fetchall()
        r = 5
        for row in result:
            ws[f"A{r}"] = row["official_use_name"]
            ws[f"P{r}"] = row["name"]
            ws[f"U{r}"] = row["date_of_acquisition"]
            r += 2

    # 表紙に記載する文章を生成する
    def gen_sentence(self, ws, department):
        sql = f"""
            SELECT department, min(official_use_name) as dept, circumstances, count(circumstances) as cnt
            FROM TW申請車両 WHERE department = '{department}' GROUP by department, circumstances;
            """
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(sql)
        result = cur.fetchall()
        cnt_list = {}
        for row in result:
            if row["circumstances"] == "A":
                cnt_list[row["circumstances"]] = row["cnt"]
            elif row["circumstances"] == "D":
                cnt_list[row["circumstances"]] = row["cnt"]
            elif row["circumstances"] == "T":
                cnt_list[row["circumstances"]] = row["cnt"]
        sent1 = f"{row['dept']}にて"
        sent2 = ""
        sent3 = ""
        sent4 = ""
        pat = ""
        if "A" in cnt_list:
            sent2 = f"{cnt_list['A']}車両増車"
            pat += "A"
        if "D" in cnt_list:
            sent3 = f"{cnt_list['D']}車両減車"
            pat += "D"
        if "T" in cnt_list:
            sent4 = f"{cnt_list['T']}車両営配"
            pat += "T"
        if pat == "A":
            sentence1 = f"{sent1}車両不足のため、{sent2}いたします。"
            sentence2 = ""
        elif pat == "D":
            sentence1 = f"{sent1}車両老朽化のため、{sent3}いたします。"
            sentence2 = ""
        elif pat == "T":
            sentence1 = f"{sent1}車両不足のため、{sent4}いたします。"
            sentence2 = ""
        elif pat == "AD":
            sentence1 = f"{sent1}車両不足のため、{sent2}いたします。"
            sentence2 = f"また、車両老朽化のため、{sent3}いたします。"
        elif pat == "AT":
            sentence1 = f"{sent1}車両不足のため、{sent2}、"
            sentence2 = f"{sent4}いたします。"
        elif pat == "DT":
            sentence1 = f"{sent1}車両老朽化のため、{sent3}いたします。"
            sentence2 = f"これにより車両不足のため、{sent4}いたします。"
        elif pat == "ADT":
            sentence1 = f"{sent1}車両不足のため、{sent2}いたします。"
            sentence2 = f"また、老朽化の{sent3}、{sent4}いたします。"
        ws["G34"] = sentence1
        ws["G35"] = sentence2

    # 管轄運輸局を取得して転記
    def get_agency(self, ws, department):
        sql = f"SELECT administrative_agency, address FROM M運輸局 WHERE code = '{department}';"
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(sql)
        result = cur.fetchone()
        ws["B4"] = result["administrative_agency"]
        ws["B5"] = result["address"]
