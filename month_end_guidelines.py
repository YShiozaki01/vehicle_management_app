import PySimpleGUI as sg
import sqlite3
import openpyxl
import os

sg.theme("SystemDefault")
DATABASE = "database/database.db"
FOLDER = "document/月末指針"
DEPTLIST = {"栃木": "02", "戸田": "03", "千葉": "04", "群馬": "05", "茨城": "07",
            "神奈川": "08", "足立": "09", "共配": "09", "名古屋": "13", "宇都宮": "22"}
COLUMNLIST1 = {1: 46, 2: 51, 3: 56, 4: 1, 5: 6, 6: 11, 7: 16, 8: 21, 9: 26, 10: 31, 11: 36, 12: 41}
COLUMNLIST2 = {1: 64, 2: 71, 3: 78, 4: 1, 5: 8, 6: 15, 7: 22, 8: 29, 9: 36, 10: 43, 11: 50, 12: 57}

# T燃料使用量の最新のidを取得
def get_new_id():
    sql_0 = "SELECT * FROM T燃料使用量;"
    sql_1 = "SELECT max(id) FROM T燃料使用量;"
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    cur.execute(sql_0)
    if len(cur.fetchall()) == 0:
        new_number = 1
    else:
        cur.execute(sql_1)
        result = cur.fetchone()
        new_number = result[0] + 1
    return new_number

    
# 「月末指針」Excelデータを取得
def get_insertdata(ws, year_num, month_num, column, dept_code):
    print(dept_code)
    i = get_new_id()
    rows = []
    c = column
    r = 3
    # 戸田以外
    if not dept_code == "03":
        while True:
            if ws.cell(r, c).value is None:
                break
            else:
                y = int(year_num) + 1 if int(month_num) < 4 else int(year_num)
                section = (i, y, int(month_num),
                        ws.cell(r, c).value,
                        dept_code,
                        0.0 if ws.cell(r, c+2).value is None else float(ws.cell(r, c+2).value),
                        0.0 if ws.cell(r, c+4).value is None else float(ws.cell(r, c+4).value))
                rows.append(section)
                i += 1
                r += 1
    # 戸田
    else:
        while True:
            if ws.cell(r, c).value is None:
                break
            else:
                y = int(year_num) + 1 if int(month_num) < 4 else int(year_num)
                section = (i, y, int(month_num),
                        ws.cell(r, c).value,
                        dept_code,
                        0.0 if ws.cell(r, c+2).value is None else float(ws.cell(r, c+2).value),
                        0.0 if ws.cell(r, c+6).value is None else float(ws.cell(r, c+6).value))
                rows.append(section)
                i += 1
                r += 1
    return rows


combo_list = []
for k in DEPTLIST.keys():
    combo_list.append(k)

layout = [
    [sg.I(k="-in1-", size=(10, 0), font=("Yu Gothic UI", 8)),
     sg.T("年度", font=("Yu Gothic UI", 11))],
    [sg.Combo([1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12], size=(3, 1), font=("Yu Gothic UI", 8),
              k="-cmb_m-", readonly=True),
     sg.T("月", font=("Yu Gothic UI", 11))],
    [sg.Combo(combo_list, size=(10, 0), font=("Yu Gothic UI", 8), k="-cmb_dept-", readonly=True)],
    [sg.B("取込み", k="-btn1-", font=("Yu Gothic UI", 8))]
]
window = sg.Window("月末指針データ取込み", layout, font=("Yu Gothic UI", 8), disable_close=False)
window.finalize()

while True:
    e, v = window.read()
    if e == "-btn1-":
        # パラメータを取得
        dept_code = DEPTLIST[v["-cmb_dept-"]]
        year_num = v["-in1-"]
        month_num = v["-cmb_m-"]
        wb_name = os.path.join(FOLDER, f"G_{year_num}.xlsx")
        ws_name = v["-cmb_dept-"]
        wb = openpyxl.load_workbook(wb_name)
        ws = wb[ws_name]
        column = COLUMNLIST1[v["-cmb_m-"]] if dept_code != "03" else COLUMNLIST2[v["-cmb_m-"]]
        # T燃料使用量テーブルに追加するデータを取得
        insert_data = get_insertdata(ws, year_num, month_num, column, dept_code)
        # T燃料使用量テーブルにデータを挿入
        sql = """
            INSERT INTO T燃料使用量 VALUES(?, ?, ?, ?, ?, ?, ?)
            """
        conn = sqlite3.connect(DATABASE)
        cur = conn.cursor()
        cur.executemany(sql, insert_data)
        conn.commit()
        # 完了メッセージ
        sg.popup(f"{ws_name}の{year_num}年{month_num}月データを追加しました。",
                 font=("Yu Gothic UI", 8), title="完了")
    if not e:
        break
window.close()
